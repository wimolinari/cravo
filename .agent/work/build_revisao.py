"""Gera revisao.xlsx refletindo o ESTADO ATUAL do site.

Duas sheets:
- Links Externos: cada par (arquivo, URL externa) com status HTTP.
- Videos YouTube: cada par (arquivo, video URL) com classificacao.

Para cada HTML em site/, varre href/src/content que comecam com http(s)://.
Para videos, parseia arrays data-videos='[...]' dos pickers.

Status HTTP e cacheado em .agent/work/link_results_full.json para acelerar
re-execucoes. Classificacao de video vem de video_classification.json (sessao
anterior) + fallback oembed para IDs novos.
"""
import sys
import json
import re
import urllib.request
import urllib.error
import ssl
import socket
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'C:\Outros\Cravo')
SITE = ROOT / 'site'
WORK = ROOT / '.agent' / 'work'
OUT = ROOT / 'revisao.xlsx'

UA = 'Mozilla/5.0 (compatible; CravoLinkChecker/1.1; revisao.xlsx)'
ATTR_RE = re.compile(r'(?:href|src|content)\s*=\s*"([^"]+)"', re.IGNORECASE)
DATA_VIDEOS_RE = re.compile(r"data-videos=\'(\[[^\']+\])\'", re.DOTALL)


def relpath(p: Path) -> str:
    return str(p.relative_to(SITE)).replace('\\', '/')


def check_url(url: str, timeout: int = 12):
    req = urllib.request.Request(url, headers={'User-Agent': UA})
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
            return resp.status, f'OK {resp.status}'
    except urllib.error.HTTPError as e:
        return e.code, f'HTTP {e.code}'
    except urllib.error.URLError as e:
        reason = getattr(e, 'reason', e)
        return None, f'UNREACHABLE: {reason}'
    except socket.timeout:
        return None, 'UNREACHABLE: timeout'
    except Exception as e:
        return None, f'UNREACHABLE: {e.__class__.__name__}'


def fetch_oembed(vid: str):
    url = f'https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={vid}&format=json'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': UA})
        with urllib.request.urlopen(req, timeout=8) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception:
        return None


def collect_link_pairs():
    pairs = []
    for p in sorted(SITE.rglob('*.html')):
        text = p.read_text(encoding='utf-8')
        origin = relpath(p)
        seen = set()
        for m in ATTR_RE.finditer(text):
            u = m.group(1).strip()
            if not (u.startswith('http://') or u.startswith('https://')):
                continue
            if u in seen:
                continue
            seen.add(u)
            pairs.append((origin, u))
    return pairs


def collect_video_pairs():
    pairs = []
    for p in sorted(SITE.rglob('*.html')):
        text = p.read_text(encoding='utf-8')
        origin = relpath(p)
        seen = set()
        for m in DATA_VIDEOS_RE.finditer(text):
            try:
                arr = json.loads(m.group(1))
            except json.JSONDecodeError:
                continue
            for item in arr:
                if not isinstance(item, dict):
                    continue
                vid = item.get('id')
                if not vid or vid in seen:
                    continue
                seen.add(vid)
                pairs.append((origin, vid))
    return pairs


def load_cache():
    p = WORK / 'link_results_full.json'
    if p.exists():
        try:
            return json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            return {}
    return {}


def save_cache(cache: dict):
    (WORK / 'link_results_full.json').write_text(
        json.dumps(cache, indent=2, ensure_ascii=False), encoding='utf-8'
    )


def main():
    print('[1/4] Coletando links externos do site...', flush=True)
    link_pairs = collect_link_pairs()
    unique = sorted({u for _, u in link_pairs})
    print(f'      {len(link_pairs)} pares, {len(unique)} URLs unicas', flush=True)

    print('[2/4] Validando HTTP (com cache em .agent/work/)...', flush=True)
    cache = load_cache()
    # Migrar entradas do link_results_v2.json se ainda nao migradas
    v2 = WORK / 'link_results_v2.json'
    if v2.exists():
        for r in json.loads(v2.read_text(encoding='utf-8')):
            u = r['url']
            if u not in cache:
                if r.get('ok'):
                    cache[u] = f'OK {r["status"]}'
                elif r.get('status'):
                    cache[u] = f'HTTP {r["status"]}'
                else:
                    cache[u] = 'UNREACHABLE (cache antigo)'
    # Gallica permanente: bloqueia minhas requisicoes mas URL e' valida (Wikipedia FR)
    cache['https://gallica.bnf.fr/ark:/12148/bpt6k1508406q'] = (
        'VERIFICADO via Wikipedia FR (Gallica bloqueia HTTP daqui)'
    )

    # Hosts usados apenas como dica de preconnect (404 na raiz e' esperado;
    # o site nao espera que o usuario clique nesses dominios).
    PRECONNECT_HINT_HOSTS = {
        'https://fonts.googleapis.com',
        'https://fonts.gstatic.com',
    }
    for u in PRECONNECT_HINT_HOSTS:
        cache[u] = 'OK (preconnect hint, nao clicavel)'

    pending = [u for u in unique if u not in cache]
    print(f'      {len(pending)} URLs a verificar agora; {len(unique) - len(pending)} ja em cache', flush=True)
    for i, u in enumerate(pending, 1):
        _, msg = check_url(u)
        cache[u] = msg
        print(f'      [{i}/{len(pending)}] {msg[:35]:35s}  {u[:90]}', flush=True)
        time.sleep(0.05)
    save_cache(cache)

    link_rows = sorted(
        ((origin, url, cache.get(url, 'NAO TESTADO')) for origin, url in link_pairs),
        key=lambda r: (r[0], r[1]),
    )

    print('[3/4] Coletando + classificando videos...', flush=True)
    video_pairs = collect_video_pairs()
    cls_data = json.loads(
        (WORK / 'video_classification.json').read_text(encoding='utf-8')
    ) if (WORK / 'video_classification.json').exists() else {}
    unique_ids = sorted({v for _, v in video_pairs})
    print(f'      {len(video_pairs)} pares, {len(unique_ids)} IDs unicos', flush=True)

    by_id = {}
    for i, vid in enumerate(unique_ids, 1):
        entry = cls_data.get(vid)
        if entry:
            cls = entry.get('classification', 'UNKNOWN')
            author = entry.get('author', '')
            reason = entry.get('reason', '')
            status = f'{cls} (canal: {author})' if author else cls
            if reason:
                status = f'{status} -- {reason}'
        else:
            oe = fetch_oembed(vid)
            if oe is None:
                status = 'UNAVAILABLE (oembed falhou; vide removido/privado?)'
            else:
                author = oe.get('author_name', '?')
                title = oe.get('title', '?')
                status = f'NOVO_NAO_CLASSIFICADO (canal: {author}, titulo: {title[:60]})'
        by_id[vid] = status

    video_rows = sorted(
        ((origin, f'https://www.youtube.com/watch?v={vid}', by_id[vid]) for origin, vid in video_pairs),
        key=lambda r: (r[0], r[1]),
    )

    print('[4/4] Gerando revisao.xlsx...', flush=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Links Externos'
    ws1.append(['Origem', 'Link', 'Status do link'])
    ws2 = wb.create_sheet('Videos YouTube')
    ws2.append(['Origem', 'Link', 'Status do link'])

    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F4F4F')
    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for r in link_rows:
        ws1.append(list(r))
    for r in video_rows:
        ws2.append(list(r))

    def color_for(s: str):
        u = s.upper()
        if u.startswith('OK ') or 'HUMAN_LIKELY' in u or 'PROBABLY_HUMAN' in u or u.startswith('VERIFICADO'):
            return 'C6EFCE'
        if 'HTTP 4' in u or 'HTTP 5' in u or 'UNAVAILABLE' in u or 'UNREACHABLE' in u:
            return 'FFC7CE'
        if 'STATIC' in u:
            return 'FFEB9C'
        if 'NOVO_NAO_CLASSIFICADO' in u:
            return 'BDD7EE'
        return None

    for ws, rows in ((ws1, link_rows), (ws2, video_rows)):
        for i, (_, _, st) in enumerate(rows, start=2):
            c = color_for(st)
            if c:
                ws.cell(row=i, column=3).fill = PatternFill('solid', fgColor=c)
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 95
        ws.column_dimensions['C'].width = 70
        ws.freeze_panes = 'A2'

    try:
        wb.save(str(OUT))
        target = OUT
    except PermissionError:
        target = OUT.with_name('revisao_atualizado.xlsx')
        print(f'      AVISO: {OUT.name} esta travado (Excel aberto?). Salvando como {target.name}', flush=True)
        wb.save(str(target))
    print(f'      OK -> {target}', flush=True)
    print(f'      Links Externos: {len(link_rows)} linhas', flush=True)
    print(f'      Videos YouTube: {len(video_rows)} linhas', flush=True)

    # Resumo no console
    from collections import Counter
    link_status_counts = Counter(s for _, _, s in link_rows)
    print('      Resumo links:', flush=True)
    for s, n in link_status_counts.most_common():
        print(f'        {n:4d}  {s}', flush=True)


if __name__ == '__main__':
    main()
