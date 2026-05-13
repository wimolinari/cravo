"""Gera VideosYoutude.xlsx (nome literal pedido pelo Wilson no PLAN.md).

Contem apenas a sheet de videos YouTube extraidos dos pickers do site.
Colunas: Origem | Link | Status do link.

Reusa cache de classificacao em .agent/work/video_classification.json.
"""
import sys
import json
import re
from pathlib import Path
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(r'C:\Outros\Cravo')
SITE = ROOT / 'site'
WORK = ROOT / '.agent' / 'work'
OUT = ROOT / 'VideosYoutude.xlsx'

DATA_VIDEOS_RE = re.compile(r"data-videos=\'(\[[^\']+\])\'", re.DOTALL)
UA = 'Mozilla/5.0 (compatible; CravoLinkChecker/1.1; VideosYoutude.xlsx)'


def relpath(p: Path) -> str:
    return str(p.relative_to(SITE)).replace('\\', '/')


def collect_video_pairs():
    pairs = []
    for p in sorted(SITE.rglob('*.html')):
        text = p.read_text(encoding='utf-8')
        origin = relpath(p)
        seen = set()
        for m in DATA_VIDEOS_RE.finditer(text):
            try:
                arr = json.loads(m.group(1))
            except json.JSONDecodeError:
                continue
            for item in arr:
                if not isinstance(item, dict):
                    continue
                vid = item.get('id')
                if not vid or vid in seen:
                    continue
                seen.add(vid)
                pairs.append((origin, vid))
    return pairs


def fetch_oembed(vid: str):
    url = f'https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={vid}&format=json'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': UA})
        with urllib.request.urlopen(req, timeout=8) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception:
        return None


def main():
    print('[1/3] Coletando videos...', flush=True)
    pairs = collect_video_pairs()
    unique_ids = sorted({v for _, v in pairs})
    print(f'      {len(pairs)} pares, {len(unique_ids)} IDs unicos', flush=True)

    print('[2/3] Classificando...', flush=True)
    cls_data = {}
    cls_file = WORK / 'video_classification.json'
    if cls_file.exists():
        cls_data = json.loads(cls_file.read_text(encoding='utf-8'))

    by_id = {}
    for vid in unique_ids:
        entry = cls_data.get(vid)
        if entry:
            cls = entry.get('classification', 'UNKNOWN')
            author = entry.get('author', '')
            reason = entry.get('reason', '')
            status = f'{cls} (canal: {author})' if author else cls
            if reason:
                status = f'{status} -- {reason}'
        else:
            oe = fetch_oembed(vid)
            if oe is None:
                status = 'UNAVAILABLE (oembed falhou; video removido/privado?)'
            else:
                author = oe.get('author_name', '?')
                title = oe.get('title', '?')
                status = f'NOVO_NAO_CLASSIFICADO (canal: {author}, titulo: {title[:60]})'
        by_id[vid] = status

    rows = sorted(
        ((origin, f'https://www.youtube.com/watch?v={vid}', by_id[vid]) for origin, vid in pairs),
        key=lambda r: (r[0], r[1]),
    )

    print('[3/3] Gerando VideosYoutude.xlsx...', flush=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Videos YouTube'
    ws.append(['Origem', 'Link', 'Status do link'])

    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F4F4F')
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for r in rows:
        ws.append(list(r))

    def color_for(s: str):
        u = s.upper()
        if 'HUMAN_LIKELY' in u or 'PROBABLY_HUMAN' in u:
            return 'C6EFCE'  # verde - humano tocando
        if 'UNAVAILABLE' in u:
            return 'FFC7CE'  # vermelho - sumido
        if 'STATIC' in u:
            return 'FFEB9C'  # amarelo - imagem estatica (alertar)
        if 'NOVO_NAO_CLASSIFICADO' in u:
            return 'BDD7EE'  # azul
        return None

    for i, (_, _, st) in enumerate(rows, start=2):
        c = color_for(st)
        if c:
            ws.cell(row=i, column=3).fill = PatternFill('solid', fgColor=c)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 95
    ws.freeze_panes = 'A2'

    try:
        wb.save(str(OUT))
        target = OUT
    except PermissionError:
        target = OUT.with_name('VideosYoutude_atualizado.xlsx')
        print(f'      AVISO: {OUT.name} travado, salvando como {target.name}', flush=True)
        wb.save(str(target))

    print(f'      OK -> {target}', flush=True)
    print(f'      {len(rows)} linhas (origem, link, status)', flush=True)

    from collections import Counter
    counts = Counter(s.split(' (')[0].split(' -')[0] for _, _, s in rows)
    print('      Resumo:', flush=True)
    for cls, n in counts.most_common():
        print(f'        {n:4d}  {cls}', flush=True)


if __name__ == '__main__':
    main()
